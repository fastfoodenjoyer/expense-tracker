"""Export transactions to Excel and Google Sheets."""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from expense_tracker.models import Transaction

logger = logging.getLogger(__name__)

# Column headers for export
EXPORT_COLUMNS = [
    "Дата",
    "Время",
    "Сумма",
    "Категория",
    "Описание",
    "Карта",
    "Банк",
]


class Exporter:
    """Export transactions to Excel and Google Sheets."""

    def __init__(
        self,
        credentials_path: Path | None = None,
        credentials_info: dict[str, Any] | None = None,
    ):
        """Initialize exporter.

        Args:
            credentials_path: Path to Google service account JSON file.
            credentials_info: Google service account info as dict (takes priority).
        """
        self.credentials_path = credentials_path
        self.credentials_info = credentials_info

    def _transaction_to_row(self, transaction: Transaction) -> list:
        """Convert transaction to a row for export."""
        return [
            transaction.date.strftime("%d.%m.%Y"),
            transaction.date.strftime("%H:%M:%S"),
            float(transaction.amount),
            transaction.category.value if transaction.category else "",
            transaction.description,
            transaction.card_number or "",
            transaction.bank,
        ]

    def export_to_excel(
        self, transactions: list[Transaction], filepath: Path
    ) -> Path:
        """Export transactions to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        # Style definitions
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, transaction in enumerate(transactions, start=2):
            row_data = self._transaction_to_row(transaction)
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths
        for col_idx, _ in enumerate(EXPORT_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(
                min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Save file
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)

        return filepath

    def _get_gspread_client(self):
        """Get authenticated gspread client.

        Supports both credentials_info dict and credentials_path file.

        Returns:
            Authenticated gspread client.

        Raises:
            ValueError: If no credentials provided.
            FileNotFoundError: If credentials file not found.
        """
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        # Prefer credentials_info (from database)
        if self.credentials_info:
            logger.info(f"Using credentials_info, client_email={self.credentials_info.get('client_email', 'N/A')}")
            credentials = Credentials.from_service_account_info(
                self.credentials_info, scopes=scopes
            )
            client = gspread.authorize(credentials)
            logger.info("gspread client authorized successfully")
            return client

        # Fallback to file path (for CLI)
        if self.credentials_path:
            if not self.credentials_path.exists():
                raise FileNotFoundError(
                    f"Google credentials not found at {self.credentials_path}"
                )
            logger.info(f"Using credentials from file: {self.credentials_path}")
            credentials = Credentials.from_service_account_file(
                str(self.credentials_path), scopes=scopes
            )
            client = gspread.authorize(credentials)
            logger.info("gspread client authorized successfully")
            return client

        raise ValueError(
            "No Google credentials provided. "
            "Set credentials via bot or provide credentials file path."
        )

    def _find_duplicates(
        self, worksheet, transactions: list[Transaction]
    ) -> set[tuple]:
        """Find existing transactions in worksheet to avoid duplicates."""
        existing_rows = set()

        all_values = worksheet.get_all_values()
        if len(all_values) > 1:
            for row in all_values[1:]:
                if len(row) >= len(EXPORT_COLUMNS):
                    existing_rows.add(tuple(row[: len(EXPORT_COLUMNS)]))

        return existing_rows

    def export_to_google_sheets(
        self,
        transactions: list[Transaction],
        spreadsheet_id: str,
        worksheet_name: str = "Транзакции",
    ) -> tuple[int, int]:
        """Export transactions to Google Sheets.

        Args:
            transactions: List of transactions to export.
            spreadsheet_id: Google Sheets spreadsheet ID.
            worksheet_name: Name of worksheet to use.

        Returns:
            Tuple of (added_count, skipped_duplicates_count).
        """
        logger.info(f"Starting export to Google Sheets: spreadsheet_id={spreadsheet_id}, transactions={len(transactions)}")

        client = self._get_gspread_client()
        logger.info(f"Opening spreadsheet by key: {spreadsheet_id}")
        spreadsheet = client.open_by_key(spreadsheet_id)
        logger.info(f"Spreadsheet opened: {spreadsheet.title}")

        # Get or create worksheet
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
            logger.info(f"Worksheet '{worksheet_name}' found")
        except Exception:
            worksheet = spreadsheet.add_worksheet(
                title=worksheet_name, rows=1000, cols=len(EXPORT_COLUMNS)
            )
            logger.info(f"Worksheet '{worksheet_name}' created")

        # Check if headers exist
        all_values = worksheet.get_all_values()
        has_correct_headers = (
            len(all_values) > 0
            and len(all_values[0]) >= len(EXPORT_COLUMNS)
            and all_values[0][:len(EXPORT_COLUMNS)] == EXPORT_COLUMNS
        )

        if not has_correct_headers:
            if not all_values:
                # Empty sheet - just add headers
                worksheet.append_row(EXPORT_COLUMNS)
                logger.info("Headers added to empty worksheet")
            else:
                # Sheet has data but no proper headers - insert at row 1
                worksheet.insert_row(EXPORT_COLUMNS, index=1)
                logger.info("Headers inserted at row 1 (existing data found without headers)")
            all_values = worksheet.get_all_values()
        else:
            logger.info("Headers already present")

        # Find duplicates
        existing_rows = self._find_duplicates(worksheet, transactions)

        # Prepare rows to add
        rows_to_add = []
        skipped = 0

        for transaction in transactions:
            row_data = self._transaction_to_row(transaction)
            row_tuple = tuple(str(v) if v != "" else "" for v in row_data)

            if row_tuple in existing_rows:
                skipped += 1
            else:
                rows_to_add.append(row_data)
                existing_rows.add(row_tuple)

        # Batch append new rows
        if rows_to_add:
            worksheet.append_rows(rows_to_add)
            logger.info(f"Appended {len(rows_to_add)} rows to worksheet")
        else:
            logger.info("No new rows to append (all duplicates)")

        logger.info(f"Export complete: added={len(rows_to_add)}, skipped={skipped}")
        return len(rows_to_add), skipped
